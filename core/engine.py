from __future__ import annotations
import io
import os
import zipfile
import xml.etree.ElementTree as ET
from typing import Any

import openpyxl


class ExportEngine:
    def __init__(self):
        self._mappers: list = []

    def register(self, mapper) -> None:
        self._mappers.append(mapper)

    @property
    def mappers(self) -> list:
        return list(self._mappers)

    def get_mapper(self, name: str):
        for m in self._mappers:
            if m.NAME == name:
                return m
        return None

    def load_file(self, path: str) -> Any:
        """Carrega um arquivo e retorna lista de dicts ou dict de abas."""
        lower = path.lower()
        if lower.endswith(".xlsx") or lower.endswith(".xls"):
            return self._load_xlsx(path)
        if lower.endswith(".xml"):
            return ET.parse(path).getroot()
        if lower.endswith(".sql"):
            with open(path, "r", encoding="utf-8", errors="replace") as f:
                return f.read()
        if lower.endswith(".json"):
            import json
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        raise ValueError(f"Formato não suportado: {path}")

    def _load_xlsx(self, path: str) -> Any:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if len(wb.sheetnames) == 1:
            return self._sheet_to_list(wb[wb.sheetnames[0]])
        return {name: self._sheet_to_list(wb[name]) for name in wb.sheetnames}

    def _sheet_to_list(self, ws) -> list[dict]:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h) if h is not None else "" for h in rows[0]]
        result = []
        for row in rows[1:]:
            result.append({headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))})
        return result

    def load_zip(self, path: str, password: str = "") -> dict[str, Any]:
        """Extrai e carrega todos os arquivos de um ZIP."""
        import json

        pwd = password.encode() if password else None
        files: dict[str, Any] = {}

        with zipfile.ZipFile(path, "r") as zf:
            for name in zf.namelist():
                try:
                    data = zf.read(name, pwd=pwd)
                except Exception:
                    continue
                lower = name.lower()
                basename = os.path.basename(name)
                if lower.endswith(".xml"):
                    try:
                        files[basename] = ET.fromstring(data.decode("utf-8", errors="replace"))
                    except Exception:
                        pass
                elif lower.endswith(".json"):
                    try:
                        files[basename] = json.loads(data.decode("utf-8"))
                    except Exception:
                        pass
                elif lower.endswith(".xlsx"):
                    try:
                        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
                        if len(wb.sheetnames) == 1:
                            files[basename] = self._sheet_to_list(wb[wb.sheetnames[0]])
                        else:
                            files[basename] = {n: self._sheet_to_list(wb[n]) for n in wb.sheetnames}
                    except Exception:
                        pass
                elif lower.endswith(".sql"):
                    try:
                        files[basename] = data.decode("utf-8", errors="replace")
                    except Exception:
                        pass
                else:
                    files[basename] = data

        return files

    def run(self, mapper_name: str, file_path: str, context: dict | None = None):
        mapper = self.get_mapper(mapper_name)
        if mapper is None:
            raise ValueError(f"Mapper '{mapper_name}' não encontrado")
        mapper.context = context or {}
        data = self.load_file(file_path)
        return mapper.extract(data)

    def run_zip(self, mapper_name: str, file_path: str, password: str = "", context: dict | None = None):
        mapper = self.get_mapper(mapper_name)
        if mapper is None:
            raise ValueError(f"Mapper '{mapper_name}' não encontrado")
        mapper.context = context or {}
        files = self.load_zip(file_path, password)
        return mapper.extract_zip(files)
